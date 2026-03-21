
from flask import Flask, render_template, request
import pandas as pd
import os
import re
from openpyxl import load_workbook

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


def get_formula_breakdown(formula):
    if isinstance(formula, str) and formula.startswith("="):
        formula = formula.replace("=", "")
        nums = re.findall(r'\d+\.?\d*', formula)
        return nums
    return []


def load_all_data():

    data = []

    files = sorted(
        os.listdir(UPLOAD_FOLDER),
        key=lambda x: os.path.getmtime(os.path.join(UPLOAD_FOLDER, x)),
        reverse=True
    )

    for file in files:

        path = os.path.join(UPLOAD_FOLDER, file)

        try:

            df = pd.read_excel(path)

            wb = load_workbook(path, data_only=False)
            ws = wb.active

            headers = [cell.value for cell in ws[1]]
            amount_col = headers.index("Total Amount") + 1

            for i, row in df.iterrows():

                supplier = str(row.get("Supplier Name", "")).strip()
                cheque = str(row.get("Cheque number", "")).replace(".0", "").strip()
                amount = row.get("Total Amount", "")

                formula = ws.cell(row=i+2, column=amount_col).value
                breakdown = get_formula_breakdown(formula)

                data.append({
                    "Department": row.get("Department", ""),
                    "Supplier Name": supplier,
                    "Total Amount": amount,
                    "Cheque number": cheque,
                    "Uploaded File": file,
                    "Breakdown": breakdown
                })

        except Exception as e:
            print("Error reading:", file, e)

    return data


@app.route("/")
def home():
    return render_template("home.html")


@app.route("/upload", methods=["GET", "POST"])
def upload():

    message = ""

    if request.method == "POST":

        file = request.files.get("file")

        if file:
            file.save(os.path.join(app.config["UPLOAD_FOLDER"], file.filename))
            message = "File uploaded successfully."

    return render_template("upload.html", message=message)


@app.route("/files")
def files():

    files = sorted(
        os.listdir(UPLOAD_FOLDER),
        key=lambda x: os.path.getmtime(os.path.join(UPLOAD_FOLDER, x)),
        reverse=True
    )

    return render_template("files.html", files=files)


@app.route("/search_supplier", methods=["GET", "POST"])
def search_supplier():

    results = []

    if request.method == "POST":

        supplier_keyword = request.form.get("supplier","").lower().strip()
        file_keyword = request.form.get("file_name","").lower().strip()

        all_data = load_all_data()

        for row in all_data:

            if supplier_keyword and supplier_keyword not in row["Supplier Name"].lower():
                continue

            if file_keyword and file_keyword not in row["Uploaded File"].lower():
                continue

            results.append(row)

    return render_template("search_supplier.html", results=results)


@app.route("/search_cheque", methods=["GET", "POST"])
def search_cheque():

    results = []

    if request.method == "POST":

        cheque_keyword = request.form.get("cheque","").strip()

        all_data = load_all_data()

        for row in all_data:

            if cheque_keyword and not row["Cheque number"].startswith(cheque_keyword):
                continue

            results.append(row)

    return render_template("search_cheque.html", results=results)


if __name__ == "__main__":
    app.run(port=3000, debug=True)
